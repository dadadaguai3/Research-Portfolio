import os
import re
import difflib
import openpyxl
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Tuple
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection

# ==============================================================================
# 1. 统一配置区 - 这里的路径已修改为相对路径，便于演示
# ==============================================================================

# --- 路径设置 ---
# 假设当前目录下有一个 data 文件夹存放原始数据
BASE_DIR = Path.cwd()
SOURCE_FOLDER = BASE_DIR / "data" / "raw_excel"
OUTPUT_FOLDER = BASE_DIR / "output"

CONSOLIDATED_FOLDER_NAME = "1_Consolidated_Files"
FINAL_PANEL_FILENAME = "专项债券面板数据汇总.xlsx"

# --- 断点续传与工作表名称配置 ---
DATA_SHEET_NAME = "Panel_Data"
LOG_SHEET_NAME = "processing_log"

# --- 阶段一 配置 (表格发现与整合) ---
# 匹配类似 "2020年...专项债券项目信息" 的表头
PATTERN_TO_FIND = r"(20(1[5-9]|2[0-5]))?.{0,3}专项债券项目信息"
SCAN_ROW_LIMIT = 3
FILENAME_KEYWORD = "专项"

# --- 阶段二、三 配置 (数据提取) ---
COLS_TO_FIND_BY_HEADER = {
    "债券名称": "债券名称",
    "发行规模": "发行规模",
}
HEADERS_TO_FIND = ["项目名称", "项目单位", "主管部门"]
DATA_COL_NAMES = ["总值", "财政安排", "债券融资"]
SPECIAL_COL_NAME = "预期总收益"
EXTENDED_VAR_NAMES = [
    "项目类型", "专项债券中用于该项目的金额", "其中：用于符合条件的重大项目资本金的金额",
    "简要描述", "建设期", "运营期", "债券存续期内项目总投资",
    "其中：不含专项债券的项目资本金", "专项债券融资", "其他债务融资"
]
# 新增的特殊查找变量
ADDITIONAL_VAR_NAME = "债券存续期内项目总收益"

# --- 阶段四 配置 (最终输出) ---
FINAL_COLUMN_ORDER = (
        list(COLS_TO_FIND_BY_HEADER.keys()) +
        HEADERS_TO_FIND +
        DATA_COL_NAMES +
        [SPECIAL_COL_NAME] +
        EXTENDED_VAR_NAMES +
        [ADDITIONAL_VAR_NAME] +
        ["年份", "地区", "数据来源"]
)


# ==============================================================================
# 2. 辅助函数区
# ==============================================================================

def clean_cell(cell_value: Any) -> str:
    """常规清理：仅去除所有空白字符。"""
    if cell_value is None or pd.isna(cell_value):
        return ""
    return re.sub(r'\s', '', str(cell_value))


def clean_for_matching(cell_value: Any) -> str:
    """强力清理：只保留中文字符，用于锚点匹配。"""
    if cell_value is None or pd.isna(cell_value):
        return ""
    # 使用正则表达式找到所有中文字符并拼接起来
    return "".join(re.findall(r'[\u4e00-\u9fa5]', str(cell_value)))


def has_border(cell: Cell) -> bool:
    """检查单元格是否有边框（用于识别表格范围）。"""
    if not cell or not cell.has_style or not cell.border:
        return False
    return any([
        cell.border.left and cell.border.left.style,
        cell.border.right and cell.border.right.style,
        cell.border.top and cell.border.top.style,
        cell.border.bottom and cell.border.bottom.style
    ])


def manual_ffill(data: List[List[Any]]) -> List[List[Any]]:
    """手动实现向下填充(Forward Fill)，用于处理合并单元格读取后的空值。"""
    if not data: return []
    filled_data = [row[:] for row in data]
    if not filled_data: return []
    num_cols = max(len(row) for row in filled_data) if filled_data else 0
    # 补齐行长度
    for row in filled_data:
        while len(row) < num_cols: row.append(None)
    # 列遍历填充
    for c in range(num_cols):
        last_val = None
        for r in range(len(filled_data)):
            if filled_data[r][c] is not None and clean_cell(filled_data[r][c]) != '':
                last_val = filled_data[r][c]
            elif last_val is not None:
                filled_data[r][c] = last_val
    return filled_data


def copy_range(source_sheet, dest_sheet, range_str, dest_start_row, dest_start_col):
    """复制指定范围的单元格（包含值和样式）到新表。"""
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    for r_idx, row in enumerate(
            source_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)):
        for c_idx, cell in enumerate(row):
            dest_cell = dest_sheet.cell(row=dest_start_row + r_idx, column=dest_start_col + c_idx)
            dest_cell.value = cell.value
            if cell.has_style:
                # 复制关键样式属性
                dest_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold,
                                      color=cell.font.color)
                dest_cell.border = Border(left=Side(style=cell.border.left.style, color=cell.border.left.color),
                                          right=Side(style=cell.border.right.style, color=cell.border.right.color),
                                          top=Side(style=cell.border.top.style, color=cell.border.top.color),
                                          bottom=Side(style=cell.border.bottom.style, color=cell.border.bottom.color))
                dest_cell.fill = PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color,
                                             end_color=cell.fill.end_color)
                dest_cell.number_format = cell.number_format
                dest_cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical,
                                                wrap_text=cell.alignment.wrap_text)

    # 处理合并单元格
    for merged_cell_range in source_sheet.merged_cells:
        if (merged_cell_range.min_row >= min_row and merged_cell_range.max_row <= max_row and
                merged_cell_range.min_col >= min_col and merged_cell_range.max_col <= max_col):
            new_min_row = dest_start_row + merged_cell_range.min_row - min_row
            new_max_row = dest_start_row + merged_cell_range.max_row - min_row
            new_min_col = dest_start_col + merged_cell_range.min_col - min_col
            new_max_col = dest_start_col + merged_cell_range.max_col - min_col
            new_range_str = f"{get_column_letter(new_min_col)}{new_min_row}:{get_column_letter(new_max_col)}{new_max_row}"
            dest_sheet.merge_cells(new_range_str)


def load_processed_log_from_excel(filepath: Path) -> set:
    """读取已处理文件的日志，实现断点续传。"""
    if not filepath.exists():
        return set()
    try:
        log_df = pd.read_excel(filepath, sheet_name=LOG_SHEET_NAME)
        processed_groups = set(log_df.iloc[:, 0].dropna().astype(str).tolist())
        print(f"检测到日志，已加载 {len(processed_groups)} 条处理记录，将跳过这些文件组。")
        return processed_groups
    except (FileNotFoundError, ValueError, KeyError):
        return set()


def append_data_and_log_to_excel(filepath: Path, new_data_df: pd.DataFrame, group_identifier: str):
    """将新提取的数据追加到Excel，并记录日志。"""
    try:
        if filepath.exists():
            with pd.ExcelFile(filepath) as xls:
                if DATA_SHEET_NAME in xls.sheet_names:
                    existing_data_df = pd.read_excel(xls, sheet_name=DATA_SHEET_NAME)
                else:
                    existing_data_df = pd.DataFrame()

                if LOG_SHEET_NAME in xls.sheet_names:
                    existing_log_df = pd.read_excel(xls, sheet_name=LOG_SHEET_NAME)
                else:
                    existing_log_df = pd.DataFrame(columns=['processed_group_id'])
        else:
            existing_data_df = pd.DataFrame()
            existing_log_df = pd.DataFrame(columns=['processed_group_id'])

        combined_data_df = pd.concat([existing_data_df, new_data_df], ignore_index=True)
        new_log_entry = pd.DataFrame({'processed_group_id': [group_identifier]})
        combined_log_df = pd.concat([existing_log_df, new_log_entry], ignore_index=True)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            combined_data_df.to_excel(writer, sheet_name=DATA_SHEET_NAME, index=False)
            combined_log_df.to_excel(writer, sheet_name=LOG_SHEET_NAME, index=False)
        print(f"  - 数据与日志已成功写入: {filepath.name}")
    except Exception as e:
        print(f"  [严重错误] 写入Excel文件 '{filepath.name}' 失败: {e}")


# ==============================================================================
# 3. 核心功能函数区
# ==============================================================================

def discover_excel_files(root_folder: Path, keyword: str) -> List[List[str]]:
    """
    扫描目录，按文件名中的编号对Excel文件进行分组。
    期望的文件名格式: name__number.xlsx
    """
    print(f"开始发现并分组Excel文件(仅限文件名包含“{keyword}”的文件)...")
    # 匹配模式：文件名__数字.xlsx
    split_file_regex = re.compile(r"(.+?)__(\d{1,3})\.xlsx$", re.IGNORECASE)
    grouped_files = {}

    for dirpath, _, filenames in os.walk(root_folder):
        for filename in filenames:
            if filename.lower().endswith('.xlsx') and keyword in filename:
                full_path = os.path.join(dirpath, filename)
                match = split_file_regex.match(filename)
                if match:
                    base_name, number = match.groups()
                    group_key = os.path.join(dirpath, base_name)
                    if group_key not in grouped_files: grouped_files[group_key] = []
                    grouped_files[group_key].append((int(number), full_path))
                else:
                    # 如果不符合 __数字 格式，作为独立文件处理
                    grouped_files[full_path] = [(0, full_path)]

    final_groups = []
    for key in sorted(grouped_files.keys()):
        file_list = sorted(grouped_files[key])
        sorted_paths = [path for number, path in file_list]
        final_groups.append(sorted_paths)
    print(f"文件发现完成，共找到 {len(final_groups)} 个符合条件的文件组。")
    return final_groups


def find_border_range_in_sheet(sheet) -> str:
    """寻找工作表中带有边框的区域（即有效数据表）。"""
    min_row, min_col, max_row, max_col = float('inf'), float('inf'), 0, 0
    found_any_border = False
    for row in sheet.iter_rows():
        for cell in row:
            if has_border(cell):
                found_any_border = True
                min_row, max_row = min(min_row, cell.row), max(max_row, cell.row)
                min_col, max_col = min(min_col, cell.column), max(max_col, cell.column)
    if found_any_border:
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    return None


def are_tables_similar(sheet1, range1, sheet2, range2) -> bool:
    """简单的表格相似度判断（基于列宽一致性）。"""
    try:
        min_col1, _, max_col1, _ = range_boundaries(range1)
        min_col2, _, max_col2, _ = range_boundaries(range2)
        return (max_col1 - min_col1 + 1) == (max_col2 - min_col2 + 1)
    except Exception:
        return False


def consolidate_tables_from_group(file_group: List[str], compiled_regex: re.Pattern, scan_row_limit: int,
                                  consolidated_output_path: Path) -> Tuple[str, List, int]:
    """
    将一组分页的Excel文件（通常是一个项目的不同页）合并为一个大的Excel表。
    """
    print(f"\n--- 正在整合文件组: {os.path.basename(file_group[0]).split('__')[0]} ---")
    all_sheets_in_group = []
    try:
        for file_path in file_group:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            all_sheets_in_group.extend(workbook.worksheets)
    except Exception as e:
        print(f"  [错误] 加载文件组时出错: {e}")
        return None, [], -1

    # 1. 找到所有包含目标关键词（如"专项债券"）的Sheet
    matched_sheets_info = []
    for index, sheet in enumerate(all_sheets_in_group):
        rows_to_scan = sheet.iter_rows(max_row=scan_row_limit) if scan_row_limit > 0 else sheet.iter_rows()
        for row in rows_to_scan:
            for cell in row:
                if cell.value and compiled_regex.search(clean_cell(cell.value)):
                    matched_sheets_info.append({'sheet': sheet, 'index': index})
                    break
            if matched_sheets_info and matched_sheets_info[-1]['sheet'] == sheet: break

    if not matched_sheets_info: return None, [], -1

    # 2. 识别连续的、结构相似的表格进行合并
    final_tables, last_sheet_index = [], -1
    for info in matched_sheets_info:
        sheet, border_range = info['sheet'], find_border_range_in_sheet(info['sheet'])
        if not border_range: continue
        final_tables.append({'sheet_object': sheet, 'range_string': border_range})
        last_sheet_index = info['index']
        next_sheet_index = info['index'] + 1

        while next_sheet_index < len(all_sheets_in_group):
            next_sheet, next_range = all_sheets_in_group[next_sheet_index], find_border_range_in_sheet(
                all_sheets_in_group[next_sheet_index])
            # 如果下一页也有边框且列宽一致，视为同一表格的续页
            if next_range and are_tables_similar(sheet, border_range, next_sheet, next_range):
                final_tables.append({'sheet_object': next_sheet, 'range_string': next_range})
                sheet, border_range = next_sheet, next_range
                last_sheet_index = next_sheet_index
                next_sheet_index += 1
            else:
                break
        break

    if not final_tables: return None, [], -1

    # 3. 执行合并写入
    processed_workbook = openpyxl.Workbook()
    dest_sheet = processed_workbook.active
    dest_sheet.title = "Consolidated_Table"
    next_row_to_write = 1
    for table_info in final_tables:
        source_sheet, range_str = table_info['sheet_object'], table_info['range_string']
        copy_range(source_sheet, dest_sheet, range_str, dest_start_row=next_row_to_write, dest_start_col=1)
        _, min_row, _, max_row = range_boundaries(range_str)
        next_row_to_write += (max_row - min_row + 1)

    try:
        output_filepath_str = str(consolidated_output_path)
        processed_workbook.save(output_filepath_str)
        return output_filepath_str, all_sheets_in_group, last_sheet_index
    except Exception as e:
        print(f"  [错误] 保存整合文件时失败: {e}")
        return None, [], -1


# --- 重构后的扩展变量提取逻辑 (核心业务逻辑) ---

def get_logical_starts_for_row(sheet, row_num: int) -> List[int]:
    """动态计算指定行的逻辑列布局（跳过合并单元格的中间列）。"""
    if not hasattr(sheet, '_row_logical_starts_cache'):
        sheet._row_logical_starts_cache = {}
    if row_num in sheet._row_logical_starts_cache:
        return sheet._row_logical_starts_cache[row_num]

    starts = []
    col = 1
    while col <= sheet.max_column:
        starts.append(col)
        # 检查是否在合并单元格内
        is_in_merge = False
        for merged_range in sheet.merged_cells.ranges:
            if (merged_range.min_row <= row_num <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                col = merged_range.max_col + 1
                is_in_merge = True
                break
        if not is_in_merge:
            col += 1

    sheet._row_logical_starts_cache[row_num] = starts
    return starts


def get_absolute_col_from_logical_index_for_row(sheet, row_num: int, logical_index: int) -> int:
    """根据指定行的逻辑索引获取绝对列号。"""
    logical_starts = get_logical_starts_for_row(sheet, row_num)
    if 1 <= logical_index <= len(logical_starts):
        return logical_starts[logical_index - 1]
    return None


def build_anchor_map(original_sheets: List, start_sheet_idx: int) -> Dict[str, List[Tuple[int, int, int]]]:
    """
    构建锚点地图：扫描后续Sheet，找到包含"项目名称"的键值对位置。
    用于从非表格区域提取扩展信息。
    """
    print("\n  [INFO] 开始构建扩展信息锚点地图...")
    anchor_map = {}

    sheets_to_scan = original_sheets[start_sheet_idx:]
    if not sheets_to_scan:
        return anchor_map

    for sheet_idx_offset, sheet in enumerate(sheets_to_scan):
        current_absolute_idx = start_sheet_idx + sheet_idx_offset

        for r_idx in range(1, sheet.max_row + 1):
            key_col = get_absolute_col_from_logical_index_for_row(sheet, r_idx, 1)
            value_col = get_absolute_col_from_logical_index_for_row(sheet, r_idx, 2)

            if not (key_col and value_col):
                continue

            key_cell_value = sheet.cell(row=r_idx, column=key_col).value

            if clean_for_matching(key_cell_value) == "项目名称":
                value_cell_value = sheet.cell(row=r_idx, column=value_col).value
                project_name = clean_for_matching(value_cell_value)

                if project_name:
                    location = (current_absolute_idx, r_idx, 2)
                    anchor_map.setdefault(project_name, []).append(location)

    print(f"  [INFO] 锚点地图构建完成。")
    return anchor_map


def extract_vars_from_anchor(anchor_location: Tuple[int, int, int], original_sheets: List) -> Dict[str, Any]:
    """从锚点位置开始，向下扫描并提取配置中定义的扩展变量。"""
    if not anchor_location:
        return {key: None for key in EXTENDED_VAR_NAMES}

    start_s_idx, start_r, value_logical_index = anchor_location
    extracted_vars = {}
    current_s_idx = start_s_idx
    current_r = start_r + 1

    # 提取常规扩展变量
    for var_name in EXTENDED_VAR_NAMES:
        found = False
        while not found and current_s_idx < len(original_sheets):
            current_sheet = original_sheets[current_s_idx]
            if current_r > current_sheet.max_row:
                current_s_idx += 1
                current_r = 1
                continue

            absolute_col = get_absolute_col_from_logical_index_for_row(current_sheet, current_r, value_logical_index)
            if not absolute_col:
                current_r += 1
                continue

            value_cell = current_sheet.cell(row=current_r, column=absolute_col)
            # 假设值在带边框的单元格中
            if has_border(value_cell):
                extracted_vars[var_name] = value_cell.value
                found = True
            current_r += 1

        if not found:
            extracted_vars[var_name] = None

    # 提取特殊变量 "债券存续期内项目总收益"
    extracted_vars[ADDITIONAL_VAR_NAME] = None
    search_limit = 10
    found_additional_var = False

    for _ in range(search_limit):
        if found_additional_var or current_s_idx >= len(original_sheets): break

        current_sheet = original_sheets[current_s_idx]
        if current_r > current_sheet.max_row:
            current_s_idx += 1
            current_r = 1
            continue

        key_col = get_absolute_col_from_logical_index_for_row(current_sheet, current_r, 1)
        value_col = get_absolute_col_from_logical_index_for_row(current_sheet, current_r, 2)

        if key_col and value_col:
            key_cell = current_sheet.cell(row=current_r, column=key_col)
            value_cell = current_sheet.cell(row=current_r, column=value_col)
            if clean_cell(key_cell.value) == ADDITIONAL_VAR_NAME and has_border(value_cell):
                extracted_vars[ADDITIONAL_VAR_NAME] = value_cell.value
                found_additional_var = True
        current_r += 1

    return extracted_vars


def extract_data_with_extended_vars(consolidated_filepath: str, original_sheets: List, last_sheet_idx: int,
                                    metadata: Dict) -> List[Dict[str, Any]]:
    """
    主提取逻辑：
    1. 从合并后的表格中读取基础数据。
    2. 利用锚点地图回到原始Sheet中查找非结构化的扩展信息。
    """
    try:
        workbook = openpyxl.load_workbook(consolidated_filepath, data_only=True)
        sheet = workbook.active
        data_original = [[cell.value for cell in row] for row in sheet.iter_rows()]
        data_filled = manual_ffill(data_original)
    except Exception as e:
        print(f"  [错误] 打开整合文件 '{os.path.basename(consolidated_filepath)}' 失败: {e}")
        return []

    # 定位表头
    header_row_idx, header_col_start_idx = None, None
    for r_idx in range(min(5, len(data_filled))):
        for c_idx in range(len(data_filled[r_idx]) - len(HEADERS_TO_FIND) + 1):
            cleaned_row_slice = [clean_cell(data_filled[r_idx][c_idx + i]) for i in range(len(HEADERS_TO_FIND))]
            if cleaned_row_slice == HEADERS_TO_FIND:
                header_row_idx, header_col_start_idx = r_idx, c_idx
                break
        if header_row_idx is not None: break

    if header_row_idx is None:
        print(f"  [警告] 在整合文件中未找到核心表头: {HEADERS_TO_FIND}。")
        return []

    # 映射表头列索引
    found_cols_map = {}
    header_row_values = [clean_cell(c) for c in data_filled[header_row_idx]]
    for final_name, header_to_find in COLS_TO_FIND_BY_HEADER.items():
        try:
            col_index = header_row_values.index(header_to_find)
            found_cols_map[final_name] = col_index
        except ValueError:
            pass

    base_extracted_data = []
    # 遍历表格行提取数据
    for i in range(header_row_idx + 1, sheet.max_row):
        proj_name = clean_cell(data_filled[i][header_col_start_idx])
        if not proj_name: continue

        row_dict = {}
        # 提取动态列
        for final_name, col_idx in found_cols_map.items():
            row_dict[final_name] = data_filled[i][col_idx] if col_idx < len(data_filled[i]) else None
        # 提取固定列
        for offset, col_name in enumerate(HEADERS_TO_FIND):
            row_dict[col_name] = data_filled[i][header_col_start_idx + offset]
        for offset, col_name in enumerate(DATA_COL_NAMES):
            col_idx = header_col_start_idx + len(HEADERS_TO_FIND) + offset
            row_dict[col_name] = data_original[i][col_idx] if col_idx < len(data_original[i]) else None

        # 提取最后一列有边框的值 (预期总收益)
        expected_return = None
        for c_idx in range(sheet.max_column, 0, -1):
            cell = sheet.cell(row=i + 1, column=c_idx)
            if has_border(cell) and cell.value is not None:
                expected_return = cell.value
                break
        row_dict[SPECIAL_COL_NAME] = expected_return
        base_extracted_data.append(row_dict)

    if not base_extracted_data: return []

    # 为项目编号以处理重名项目
    name_counts = {}
    for record in base_extracted_data:
        cleaned_name = clean_for_matching(record.get("项目名称"))
        if cleaned_name:
            current_count = name_counts.get(cleaned_name, 0) + 1
            record['_occurrence_index'] = current_count
            name_counts[cleaned_name] = current_count

    # 构建并使用锚点地图提取扩展信息
    anchor_map = build_anchor_map(original_sheets, last_sheet_idx + 1)
    fully_enriched_data = []

    for record in base_extracted_data:
        project_name = record.get("项目名称")
        cleaned_project_name = clean_for_matching(project_name)
        occurrence_index = record.get('_occurrence_index', -1)

        anchor_locations = anchor_map.get(cleaned_project_name, [])

        if 1 <= occurrence_index <= len(anchor_locations):
            anchor_location = anchor_locations[occurrence_index - 1]
            extended_vars = extract_vars_from_anchor(anchor_location, original_sheets)
        else:
            extended_vars = {key: None for key in EXTENDED_VAR_NAMES}
            extended_vars[ADDITIONAL_VAR_NAME] = None

        record.update(extended_vars)
        if '_occurrence_index' in record: del record['_occurrence_index']
        record.update(metadata)
        fully_enriched_data.append(record)

    return fully_enriched_data


# ==============================================================================
# 4. 主程序区
# ==============================================================================
def main():
    """主程序，负责调度所有模块。"""

    # 初始化输出目录
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    consolidated_dir = OUTPUT_FOLDER / CONSOLIDATED_FOLDER_NAME
    consolidated_dir.mkdir(exist_ok=True)
    final_output_path = OUTPUT_FOLDER / FINAL_PANEL_FILENAME

    # 加载断点日志
    processed_groups = load_processed_log_from_excel(final_output_path)

    compiled_regex = re.compile(PATTERN_TO_FIND, re.IGNORECASE)

    # 检查输入源
    if not SOURCE_FOLDER.exists():
        print(f"[错误] 数据源目录不存在: {SOURCE_FOLDER}")
        print("请在代码中配置正确的 SOURCE_FOLDER 或创建相应目录。")
        return

    excel_file_groups = discover_excel_files(SOURCE_FOLDER, FILENAME_KEYWORD)
    if not excel_file_groups:
        print("在指定目录中未发现任何符合条件的文件。")
        return

    print(f"\n>>>>>> 开始处理 {len(excel_file_groups)} 个文件组 <<<<<<")

    for i, file_group in enumerate(excel_file_groups, 1):
        group_identifier = file_group[0]
        base_name_for_output = os.path.basename(group_identifier).split('__')[0]
        print(f"\n[{i}/{len(excel_file_groups)}] 正在处理: {base_name_for_output}")

        if group_identifier in processed_groups:
            print("  - 跳过: 已在日志中标记为完成。")
            continue

        consolidated_output_path = consolidated_dir / f"{base_name_for_output}_consolidated.xlsx"

        # 步骤 1: 整合表格
        consolidated_filepath, original_sheets, last_sheet_idx = consolidate_tables_from_group(
            file_group, compiled_regex, SCAN_ROW_LIMIT, consolidated_output_path
        )

        if not consolidated_filepath:
            print(f"  - 失败: 表格整合失败，跳过。")
            continue

        # 提取元数据 (注意：此处假设了文件夹结构为 地区/年份/文件)
        # 实际使用时请根据自己的文件夹层级调整
        try:
            year = Path(file_group[0]).parent.name
            region = Path(file_group[0]).parent.parent.name
        except IndexError:
            year, region = "未知", "未知"

        metadata = {"年份": year, "地区": region, "数据来源": str(Path(group_identifier).name)}

        # 步骤 2: 提取数据
        enriched_data = extract_data_with_extended_vars(
            consolidated_filepath, original_sheets, last_sheet_idx, metadata
        )

        # 步骤 3: 写入结果
        if enriched_data:
            new_data_df = pd.DataFrame(enriched_data)
            # 补全缺失列
            for col in FINAL_COLUMN_ORDER:
                if col not in new_data_df.columns:
                    new_data_df[col] = None
            new_data_df = new_data_df[FINAL_COLUMN_ORDER]

            append_data_and_log_to_excel(final_output_path, new_data_df, group_identifier)
        else:
            print(f"  - 警告: 未提取到有效数据。")

    print(f"\n================== [全部任务完成] ==================")
    print(f"输出文件: {final_output_path.resolve()}")


if __name__ == "__main__":
    main()